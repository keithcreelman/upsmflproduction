#!/usr/bin/env python3
"""
Extract extensions from the 2019/2020/2021/undated Contract Transaction Log
xlsx files and emit SQL ready to UPSERT into ups_extension_master.

Sources (in pipelines/etl/inputs/):
  • 2019_Contract_Tansaction_Log.xlsx
  • 2020 Contract_Transaction_Log.xlsx
  • 2021_Contract_Transaction_Log.xlsx
  • Contract_Transaction_Log.xlsx   (undated — Keith to confirm year)

For each file we read the "Extensions" sheet and:
  1. Parse header to locate columns (offsets differ between files).
  2. Map "Extension By" team-name → franchise_id via src_franchises.
  3. Strip "Last, First TEAM POS" → "Last, First" and resolve player_id
     via player_id_crosswalk (full_name).
  4. Emit one UPSERT per (season, player_id) with evidence_grade='evidenced'
     and evidence_source='contract_transaction_log:<filename>'.

Dedup strategy:
  • Within a file: keep the LAST row per (season, player_id) — the file is
    chronologically ordered so a later edit wins.
  • Across files: the SQL is UPSERT (ON CONFLICT ... DO UPDATE), so the
    INSERT order in the final .sql matters. We process files in this order:
      2019, 2020, 2021, undated
    which means undated wins if it touches the same (season, player_id).

Cross-check:
  • The Pre-Extension Contract column gives "current year" salary; the
    salary-year columns give post-extension salaries. Term derives from
    "Years Extended" col directly — Keith's canonical: EXT1=1, EXT2=2.
  • Owner short-names ("Rico" = first name of Rico Balderelli, fr=0005).

Output: writes /tmp/extensions_2019_2021_upserts.sql (review before applying).
"""

from __future__ import annotations

import os
import re
import sys
import warnings
from collections import defaultdict
from pathlib import Path

import openpyxl

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

REPO_ROOT = Path('/Users/keithcreelman/Code/MFL/upsmflproduction')
INPUTS = REPO_ROOT / 'pipelines/etl/inputs'

FILES = [
    ('2019', '2019_Contract_Tansaction_Log.xlsx'),
    ('2020', '2020 Contract_Transaction_Log.xlsx'),
    ('2021', '2021_Contract_Transaction_Log.xlsx'),
    # Keith confirmed 2026-05-24: "This has 2022 in excel file" — the
    # undated Contract_Transaction_Log.xlsx covers 2022 extensions.
    ('2022', 'Contract_Transaction_Log.xlsx'),
]

# Team-name → franchise_id (stable 2019-2021 per src_franchises). Includes
# common shorthands seen in the xlsx "Extension By" column.
TEAM_TO_FID = {
    '#blm': '0008',
    'blm': '0008',
    'keith': '0008',
    'the bash bros': '0002',
    'bash bros': '0002',
    'aj': '0002',
    'gride': '0003',
    'matt': '0003',
    'pure greatness': '0004',
    'brian': '0004',
    'run cmc': '0005',
    'rico': '0005',
    'good in da hood': '0006',
    'steve': '0006',
    'sex manther': '0007',
    'josh': '0007',
    'c-town chivalry': '0009',
    'ctown chivalry': '0009',
    'c-town': '0009',
    'bear': '0009',
    'blake bombers': '0010',
    'shawn': '0010',
    'cleon ca$h': '0011',
    'cleon cash': '0011',
    'eric': '0011',
    'hawks': '0012',
    'chris': '0012',
    # Stale team name from 2019 (Mike Williams ext): Keith confirms Hawks.
    'white power': '0012',
    'ulterior warrior': '0001',
    'ryan': '0001',
}


def norm_team(s: str) -> str:
    return re.sub(r'\s+', ' ', (s or '').strip().lower())


def resolve_franchise_id(extension_by: str) -> str | None:
    key = norm_team(extension_by)
    if key in TEAM_TO_FID:
        return TEAM_TO_FID[key]
    # try first token
    first = key.split()[0] if key else ''
    if first in TEAM_TO_FID:
        return TEAM_TO_FID[first]
    return None


def normalize_player_name(raw: str) -> str:
    """'McCaffrey, Christian CAR RB' → 'Christian McCaffrey' (for crosswalk lookup).
    The crosswalk uses 'full_name' which appears to be 'First Last' format.
    """
    if not raw:
        return ''
    s = str(raw).strip()
    # Strip trailing " TEAM POS" pattern (uppercase 2-4 letter team + 1-3 letter pos)
    s = re.sub(r'\s+[A-Z]{2,4}\s+[A-Z]{1,3}$', '', s)
    if ',' in s:
        last, first = s.split(',', 1)
        s = f"{first.strip()} {last.strip()}"
    return ' '.join(s.split())


def extract_nfl_team(raw: str) -> str:
    """'McCaffrey, Christian CAR RB' → 'CAR'. Returns '' if not parseable."""
    if not raw:
        return ''
    m = re.search(r'\s+([A-Z]{2,4})\s+[A-Z]{1,3}$', str(raw).strip())
    return m.group(1) if m else ''


SEASON_BY_LABEL = {'2019': 2019, '2020': 2020, '2021': 2021, '2022': 2022}


def extract_extensions(filepath: Path, file_label: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if 'Extensions' not in wb.sheetnames:
        wb.close()
        return []
    ws = wb['Extensions']
    # Parse header at row 3 to locate columns
    header = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None and str(h).strip()}

    def col(*names) -> int | None:
        for n in names:
            if n.lower() in idx:
                return idx[n.lower()] + 1  # 1-indexed
        return None

    c_ext_by = col('extension by:', 'extension by')
    c_player = col('player')
    c_years = col('years extended')
    c_type = col('type extension')
    c_pos = col('position')
    c_pre = col('pre-extension contract')
    c_rookie = col('expired rookie')
    c_inc = col('salary increase')
    c_notes = col('notes')
    c_ts = col('timestamp', 'time stamp')

    if not all([c_ext_by, c_player, c_years]):
        print(f'  ! could not locate required columns in {filepath.name}', file=sys.stderr)
        wb.close()
        return []

    out = []
    for ri in range(4, ws.max_row + 1):
        ext_by = ws.cell(row=ri, column=c_ext_by).value
        player_raw = ws.cell(row=ri, column=c_player).value
        years = ws.cell(row=ri, column=c_years).value
        if not (ext_by or player_raw):
            continue
        if not player_raw:
            continue
        try:
            term = int(years) if years is not None else None
        except (TypeError, ValueError):
            term = None
        ext_type = (ws.cell(row=ri, column=c_type).value or '') if c_type else ''
        position = (ws.cell(row=ri, column=c_pos).value or '') if c_pos else ''
        pre_salary = ws.cell(row=ri, column=c_pre).value if c_pre else None
        expired_rookie = ws.cell(row=ri, column=c_rookie).value if c_rookie else None
        new_aav = ws.cell(row=ri, column=c_inc).value if c_inc else None
        notes = (ws.cell(row=ri, column=c_notes).value or '') if c_notes else ''
        ts = ws.cell(row=ri, column=c_ts).value if c_ts else None

        out.append({
            'extension_by_raw': str(ext_by or '').strip(),
            'player_raw': str(player_raw).strip(),
            'player_name': normalize_player_name(str(player_raw)),
            'nfl_team': extract_nfl_team(str(player_raw)),
            'term': term,
            'type_extension': str(ext_type).strip(),
            'position': str(position).strip().upper(),
            'pre_salary': pre_salary,
            'expired_rookie': bool(expired_rookie),
            'salary_increase': new_aav,
            'notes': str(notes).strip(),
            'timestamp': ts.isoformat() if hasattr(ts, 'isoformat') else (str(ts) if ts else None),
            'source_file': filepath.name,
            'file_label': file_label,
        })
    wb.close()
    return out


def main():
    all_rows = []
    for label, fn in FILES:
        path = INPUTS / fn
        if not path.exists():
            print(f'! missing: {path}', file=sys.stderr)
            continue
        rows = extract_extensions(path, label)
        print(f'{label:>8} {fn:<45} rows={len(rows)}', file=sys.stderr)
        all_rows.extend(rows)

    # Resolve franchise_id + emit summary
    unmatched_team = defaultdict(int)
    for r in all_rows:
        fid = resolve_franchise_id(r['extension_by_raw'])
        r['franchise_id'] = fid
        if not fid:
            unmatched_team[r['extension_by_raw']] += 1

    if unmatched_team:
        print('\nUnmatched team names (need TEAM_TO_FID entry):', file=sys.stderr)
        for k, n in sorted(unmatched_team.items(), key=lambda x: -x[1]):
            print(f'  {k!r}: {n}', file=sys.stderr)

    # Resolve player_ids via the player_id_crosswalk D1 table.
    # We pull the full crosswalk once and match by normalized full_name.
    import json
    import subprocess

    def fetch_crosswalk() -> dict[str, list[dict]]:
        """Build name-normalized → list of {id, position, nfl_team, season}
        lookup from D1 src_players (canonical MFL TYPE=players mirror).

        Per-season rows mean the same player_id appears multiple times
        with potentially different nfl_team. We keep ALL season variants
        so disambiguation can match on NFL team for the right year.
        """
        idx = defaultdict(list)
        cmd = [
            'npx', 'wrangler', 'd1', 'execute', 'ups-mfl-db', '--remote',
            '--command',
            'SELECT season, player_id, name, position, nfl_team FROM src_players;',
            '--json',
        ]
        cp = subprocess.run(cmd, cwd=str(REPO_ROOT / '.claude/worktrees/keen-knuth-623b0f/worker'),
                            capture_output=True, text=True, check=False)
        if cp.returncode != 0:
            print(f'! src_players fetch failed: {cp.stderr[:300]}', file=sys.stderr)
            return idx
        data = json.loads(cp.stdout)
        rows = data[0]['results'] if isinstance(data, list) else data.get('results', [])
        for row in rows:
            nm = row.get('name') or ''   # MFL convention: 'Last, First'
            if not nm:
                continue
            # Build lookup key from BOTH "Last, First" and "First Last" forms
            if ',' in nm:
                last, first = nm.split(',', 1)
                nm_norm = f"{first.strip()} {last.strip()}"
            else:
                nm_norm = nm
            key = re.sub(r'[^a-z]', '', nm_norm.lower())
            if not key:
                continue
            idx[key].append({
                'id': str(row['player_id']),
                'position': row.get('position', '') or '',
                'nfl_team': row.get('nfl_team', '') or '',
                'season': int(row['season']) if row.get('season') else 0,
                'full_name': nm_norm,
                'source': 'src_players',
            })
        return idx

    print('\nResolving player_ids from crosswalk...', file=sys.stderr)
    xwalk = fetch_crosswalk()
    print(f'  crosswalk entries: {sum(len(v) for v in xwalk.values())}', file=sys.stderr)

    matched = 0
    unmatched = []
    for r in all_rows:
        nm = r['player_name']
        key = re.sub(r'[^a-z]', '', nm.lower())
        candidates = xwalk.get(key, [])
        season = SEASON_BY_LABEL.get(r['file_label'], 0)
        pos = (r.get('position') or '').upper()
        team = (r.get('nfl_team') or '').upper()

        # Disambiguate via cascading filters:
        #   1. exact (position + nfl_team + season)
        #   2. exact (position + nfl_team)   — any season
        #   3. (position + season)
        #   4. position only
        #   5. unique by id (single player_id across multiple season rows)
        def with_filter(*preds):
            return [c for c in candidates if all(p(c) for p in preds)]

        winner_ids = None
        for filters in (
            (lambda c: c['position'].upper() == pos,
             lambda c: c['nfl_team'].upper() == team,
             lambda c: c['season'] == season),
            (lambda c: c['position'].upper() == pos,
             lambda c: c['nfl_team'].upper() == team),
            (lambda c: c['position'].upper() == pos,
             lambda c: c['season'] == season),
            (lambda c: c['position'].upper() == pos,),
        ):
            best = with_filter(*filters)
            unique = {c['id'] for c in best}
            if len(unique) == 1:
                winner_ids = unique
                break

        if winner_ids is None:
            # Fall back: even without position match, see if all crosswalk
            # entries for this name resolve to a single player_id.
            unique = {c['id'] for c in candidates}
            if len(unique) == 1:
                winner_ids = unique

        if winner_ids:
            r['player_id'] = next(iter(winner_ids))
            matched += 1
        else:
            r['player_id'] = None
            unique = {c['id'] for c in candidates}
            unmatched.append((nm, pos, team, len(unique)))

    print(f'  matched: {matched} / {len(all_rows)}', file=sys.stderr)
    if unmatched:
        print(f'\nUnmatched players ({len(unmatched)}):', file=sys.stderr)
        for nm, pos, team, n in unmatched[:30]:
            print(f'  {nm} [{pos} {team}] distinct_ids={n}', file=sys.stderr)

    # Emit CSV of rows for review
    out_csv = Path('/tmp/extensions_2019_2021_raw.csv')
    import csv
    with out_csv.open('w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=[
            'file_label', 'source_file', 'franchise_id', 'extension_by_raw',
            'player_id', 'player_name', 'player_raw', 'position', 'term',
            'type_extension', 'pre_salary', 'expired_rookie',
            'salary_increase', 'notes', 'timestamp',
        ])
        w.writeheader()
        for r in all_rows:
            w.writerow({k: r.get(k) for k in w.fieldnames})
    print(f'\nWrote {len(all_rows)} rows to {out_csv}', file=sys.stderr)

    # Emit SQL UPSERTs for resolved rows only. Unresolved (no player_id
    # or no franchise_id) are skipped — they go to parking_lot manually
    # or via a follow-up after Keith reviews unmatched names.
    out_sql = Path('/tmp/extensions_2019_2021_upserts.sql')
    sql_rows = [r for r in all_rows
                if r.get('player_id') and r.get('franchise_id')
                and r.get('term') in (1, 2)]
    # Dedup within-file: keep LAST row per (season, player_id) seen
    keep = {}
    for r in sql_rows:
        keep[(r['file_label'], r['player_id'])] = r
    sql_rows = list(keep.values())

    def sql_str(v):
        if v is None:
            return 'NULL'
        s = str(v).replace("'", "''")
        return f"'{s}'"

    def sql_int(v):
        try:
            return str(int(v))
        except (TypeError, ValueError):
            return 'NULL'

    with out_sql.open('w') as f:
        f.write("""-- 0062_extension_master_2019_2021_xlsx_backfill.sql
-- 2019-2021 extension backfill sourced from Keith's Contract Transaction
-- Log spreadsheets (xlsx attached 2026-05-24). Evidence grade='evidenced'
-- because each row came from the canonical league-tracked Google Form
-- response sheet; evidence_source identifies the file.
--
-- Skipped:
--   • Contract_Transaction_Log.xlsx (undated) — Keith asked to skip; only
--     10/34 player overlap with 2021 file, salary cols 2020/2021/2022 are
--     ambiguous. Handle in a follow-up after scope is clear.
--   • Rows with unresolved player_id or franchise_id (logged in /tmp/
--     extensions_2019_2021_raw.csv for manual review).
--
-- Dedup: within-file last-wins by (season, player_id).
--

""")
        for r in sorted(sql_rows, key=lambda x: (x['file_label'], x['franchise_id'], x['player_name'])):
            status = 'EXT1' if r['term'] == 1 else 'EXT2'
            extended_at = r.get('timestamp') or f"{r['file_label']}-01-01T00:00:00Z"
            evidence_src = f"contract_transaction_log:{r['source_file']}"
            try:
                pre_sal_int = int(r['pre_salary']) if r.get('pre_salary') is not None else None
            except (TypeError, ValueError):
                pre_sal_int = None
            try:
                inc_int = int(r['salary_increase']) if r.get('salary_increase') is not None else None
            except (TypeError, ValueError):
                inc_int = None
            # Canon §C4: extension-year AAV = pre-extension salary + position
            # bump (Sched 1: +10K/1yr, +20K/2yr; Sched 2: +3K/1yr, +5K/2yr).
            # The xlsx "Salary Increase" column stores the BUMP amount,
            # not the post-extension AAV. Compute post-ext AAV correctly.
            #
            # TCV = current year (pre-ext salary, carries) + extension
            # years (at new AAV). Canon §C4 worked example:
            #   1yr remaining $17K + ext 1yr → TCV = 17K + 27K = 44K
            #   Not 17K + 10K = 27K, which is what the bug computed.
            #
            # For expired-rookie extensions (no carry, fresh contract),
            # ALL years are at new AAV: TCV = new_aav * term.
            # Heuristic: if pre_salary <= 0 OR expired_rookie flag set,
            # treat as fresh contract.
            new_aav = (pre_sal_int + inc_int) if (pre_sal_int is not None and inc_int is not None) else inc_int
            new_salary = new_aav  # year-1 of the new deal
            if r.get('expired_rookie') or not pre_sal_int:
                # Fresh contract — all years at new AAV
                new_tcv = (new_aav * r['term']) if (new_aav and r['term']) else None
                # year-1 of a fresh contract is the new AAV
                new_salary = new_aav
            elif new_aav is not None and pre_sal_int is not None:
                # Vet extension — current year carries at pre-ext salary,
                # extension years at new AAV
                new_tcv = pre_sal_int + (new_aav * r['term'])
                # Y1 (current year) stays at pre-ext salary
                new_salary = pre_sal_int
            else:
                new_tcv = None
            new_gtd = round(new_tcv * 0.75) if new_tcv else None

            f.write(f"""INSERT INTO ups_extension_master
  (league_id, season, franchise_id, player_id, player_name, position,
   new_contract_status, new_salary, new_contract_year, new_contract_info,
   extension_term_years, new_tcv, new_aav, new_gtd, ext_token,
   source, extended_at_utc, updated_at_utc,
   evidence_grade, evidence_source)
VALUES
  ('74598', {sql_str(r['file_label'])}, {sql_str(r['franchise_id'])}, {sql_str(r['player_id'])},
   {sql_str(r['player_name'])}, {sql_str(r['position'])},
   {sql_str(status)}, {sql_int(new_salary)}, {sql_int(r['term'])}, NULL,
   {sql_int(r['term'])}, {sql_int(new_tcv)}, {sql_int(new_aav)}, {sql_int(new_gtd)}, NULL,
   'contract-transaction-log-xlsx', {sql_str(extended_at)}, datetime('now'),
   'evidenced', {sql_str(evidence_src)})
ON CONFLICT(league_id, season, player_id) DO UPDATE SET
  franchise_id         = excluded.franchise_id,
  player_name          = COALESCE(NULLIF(excluded.player_name, ''), ups_extension_master.player_name),
  position             = COALESCE(NULLIF(excluded.position, ''), ups_extension_master.position),
  new_contract_status  = excluded.new_contract_status,
  new_salary           = COALESCE(excluded.new_salary, ups_extension_master.new_salary),
  new_contract_year    = excluded.new_contract_year,
  extension_term_years = excluded.extension_term_years,
  new_tcv              = COALESCE(excluded.new_tcv, ups_extension_master.new_tcv),
  new_aav              = COALESCE(excluded.new_aav, ups_extension_master.new_aav),
  new_gtd              = COALESCE(excluded.new_gtd, ups_extension_master.new_gtd),
  source               = excluded.source,
  updated_at_utc       = excluded.updated_at_utc,
  evidence_grade       = excluded.evidence_grade,
  evidence_source      = excluded.evidence_source;

""")

    print(f'\nWrote {len(sql_rows)} INSERT statements to {out_sql}', file=sys.stderr)


if __name__ == '__main__':
    main()
